import json
import os
import sys
from collections import defaultdict

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill


def main():
    filename = sys.argv[1]

    data = defaultdict(lambda: {
        "result": "",
        "reasoning": ""
    })

    # Parse JSON-Objekte aus der Datei
    with open(filename, 'r', encoding='utf-8') as f:
        json_objects = [json.loads(line) for line in f]

    print(f"Gefundene JSON-Objekte: {len(json_objects)}")

    for entry in json_objects:
        einrichtung = entry.get("einrichtung", "")
        result = entry.get("result", False)
        reasoning = entry.get("reasoning", "")
        print(f"Processing entry: {einrichtung}")

        if einrichtung:
            data[einrichtung]["result"] = "yes" if result else "no"
            data[einrichtung]["reasoning"] = reasoning

    print("Gefundene Einrichtungen:", len(data))

    # Excel-Export
    excel_filename = os.path.splitext(filename)[0] + "_report.xlsx"
    print(f"Erstelle Excel-Bericht: {excel_filename}")
    create_excel_report(data, excel_filename)


def create_excel_report(data, output_filename: str):
    """Erstellt eine Excel-Tabelle mit Einrichtungen als Zeilen"""

    with_reasoning = False

    # Erstelle DataFrame
    rows = []
    for einrichtung in sorted(data.keys()):
        row = {"Einrichtung": einrichtung}

        row["result"] = data[einrichtung]["result"]
        if with_reasoning:
            row["result_Begründung"] = data[einrichtung]["reasoning"]

        rows.append(row)

    df = pd.DataFrame(rows)

    # Speichere als Excel-Datei
    df.to_excel(output_filename, index=False, engine='openpyxl')
    print(f"Excel-Bericht erstellt: {output_filename}")

    # Formatierung
    with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a') as writer:
        worksheet = writer.sheets['Sheet1']

        # Setze Spaltenbreiten
        worksheet.column_dimensions['A'].width = 76
        if with_reasoning:
            worksheet.column_dimensions['C'].width = 30

        # Setze Überschriften fett
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Standard Excel palette colors
        green_fill = PatternFill(
            start_color="ceeed0", end_color="ceeed0", fill_type="solid")  # Light green
        red_fill = PatternFill(
            start_color="f6c9ce", end_color="f6c9ce", fill_type="solid")  # Light red
        green_text = Font(color="285f17")  # Dark green text
        red_text = Font(color="8f1b15")  # Dark red text

        area = f"B2:B{len(df) + 1}"

        rule = CellIsRule(operator="equal", formula=[
                          '"yes"'], stopIfTrue=True, fill=green_fill, font=green_text)
        worksheet.conditional_formatting.add(area, rule)

        rule = CellIsRule(operator="equal", formula=[
                          '"no"'], stopIfTrue=True, fill=red_fill, font=red_text)
        worksheet.conditional_formatting.add(area, rule)


if __name__ == "__main__":
    main()
