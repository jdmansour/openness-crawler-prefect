import sys
import os
import json
import pandas as pd
from collections import defaultdict

from openpyxl.formatting import Rule
from openpyxl.formatting.rule import CellIsRule

from openpyxl.styles import Font, PatternFill, Border
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.styles.differential import DifferentialStyle

def main():
    filename = sys.argv[1]

    data = defaultdict(lambda: defaultdict(lambda: {
        "result": "",
        "reasoning": ""
    }))

    # Parse JSON-Objekte aus der Datei
    with open(filename, 'r', encoding='utf-8') as f:
        json_objects = [json.loads(line) for line in f]

    print(f"Gefundene JSON-Objekte: {len(json_objects)}")
    
    for entry in json_objects:
        einrichtung = entry.get("einrichtung", "")
        software = entry.get("software", "")
        result = entry.get("result", False)
        reasoning = entry.get("reasoning", "")
        print(f"Processing entry: {einrichtung}, {software}")

        if einrichtung and software:
            data[einrichtung][software]["result"] = "yes" if result else "no"
            data[einrichtung][software]["reasoning"] = reasoning

    print("Gefundene Einrichtungen:", len(data))

    # for item in list(data.items())[:5]:
    #     einrichtung, software_data = item
    #     print(f"Einrichtung: {einrichtung}")
    #     for software, details in software_data.items():
    #         print(f"  Software: {software}, Nutzung: {details['result']}, Begründung: {details['reasoning']}")
    #     print()

    # Excel-Export
    excel_filename = os.path.splitext(filename)[0] + "_report.xlsx"
    print(f"Erstelle Excel-Bericht: {excel_filename}")
    create_excel_report(data, excel_filename)

def create_excel_report(data, output_filename: str):
    """Erstellt eine Excel-Tabelle mit Einrichtungen als Zeilen und Software-Programmen als Spalten"""
    
    # Sammle alle einzigartigen Software-Programme
    all_software = set()
    for einrichtung_data in data.values():
        all_software.update(einrichtung_data.keys())
    
    all_software = sorted(list(all_software))

    with_reasoning = False
    
    # Erstelle DataFrame
    rows = []
    for einrichtung in sorted(data.keys()):
        row = {"Einrichtung": einrichtung}
        
        # Für jede Software füge sowohl Nutzung als auch Begründung hinzu
        for software in all_software:
            if software in data[einrichtung]:
                row[f"{software}"] = data[einrichtung][software]["result"]
                if with_reasoning:
                    row[f"{software}_Begründung"] = data[einrichtung][software]["reasoning"]
            else:
                row[f"{software}"] = "no"
                if with_reasoning:
                    row[f"{software}_Begründung"] = "(No usage found in any document)"
        
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    # Speichere als Excel-Datei
    df.to_excel(output_filename, index=False, engine='openpyxl')
    print(f"Excel-Bericht erstellt: {output_filename}")
    

    # Formatiere die Tabelle:
    # Spalte A: Breite 76
    # Spalten B, E, G: Breite 30
    with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Setze Spaltenbreiten
        worksheet.column_dimensions['A'].width = 76
        if with_reasoning:
            worksheet.column_dimensions['C'].width = 30
            worksheet.column_dimensions['E'].width = 30
            worksheet.column_dimensions['G'].width = 30
        
        # Setze Überschriften fett
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Bedingte formatierung für Spalten B, D, F:
        # Wenn der inhalt yes ist, hinterlege die Zelle grün
        # Wenn der Inhalt no ist, hinterlege die Zelle rot

        #for col in ['B', 'D', 'F']:
    
        # Standard Excel palette colors (indexed)
        green_fill = PatternFill(start_color="ceeed0", end_color="ceeed0", fill_type="solid")  # Light green
        red_fill   = PatternFill(start_color="f6c9ce", end_color="f6c9ce", fill_type="solid")  # Light red
        red_text = Font(color="8f1b15")  # Dark red text
        green_text = Font(color="285f17")  # Dark green text
        # Area is the whole of B, D, F:
        n = len(df) + 1
        if with_reasoning:
            area = f"B2:B{n} D2:D{n} F2:F{n}"
        else:
            area = f"B2:B{n} C2:C{n} D2:D{n}"

        # for col in ['B', 'D', 'F']:
            # rule = Rule(
            #     type="containsText",
            #     operator="containsText",
            #     text="yes",
            #     dxf=DifferentialStyle(
            #         fill=PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            #     )
            # )

        rule = CellIsRule(operator="equal", formula=['"yes"'], stopIfTrue=True, fill=green_fill, font=green_text)
        worksheet.conditional_formatting.add(area, rule)

        rule = CellIsRule(operator="equal", formula=['"no"'], stopIfTrue=True, fill=red_fill, font=red_text)
        worksheet.conditional_formatting.add(area, rule)
        # workbook.save(output_filename)


    # # Zeige eine Vorschau der ersten paar Zeilen
    # print("\nVorschau der ersten 3 Zeilen:")
    # print(df.head(3).to_string(index=False))

if __name__ == "__main__":
    main()